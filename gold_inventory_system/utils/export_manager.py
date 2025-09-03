from __future__ import annotations

import os
import tempfile
from datetime import datetime
from typing import List, Dict

LIGHT_GREEN = "#90EE90"
LIGHT_RED = "#FFB6C1"


class ExportManager:
    @staticmethod
    def _html_template(title: str, table_html: str, subtitle: str) -> str:
        return f"""
<!DOCTYPE html>
<html lang=\"ar\" dir=\"rtl\">
<head>
  <meta charset=\"utf-8\" />
  <title>{title}</title>
  <style>
    body {{ font-family: 'Noto Naskh Arabic', sans-serif; background: #fafafa; color: #222; }}
    .container {{ width: 960px; margin: 30px auto; background: #fff; padding: 20px 24px; border: 1px solid #eee; }}
    h1 {{ text-align: center; margin: 0 0 8px; }}
    .subtitle {{ text-align: center; margin-bottom: 20px; color: #555; }}
    table {{ width: 100%; border-collapse: collapse; }}
    th, td {{ border: 1px solid #e0e0e0; padding: 8px 10px; }}
    th {{ background: #f5f5f5; }}
    .inc {{ background: {LIGHT_GREEN}; }}
    .dec {{ background: {LIGHT_RED}; }}
  </style>
  <meta name=\"viewport\" content=\"width=device-width, initial-scale=1\" />
  <meta name=\"author\" content=\"مصنع الذهب\" />
  <meta name=\"generator\" content=\"Gold Inventory System\" />
  <meta name=\"date\" content=\"{datetime.now().isoformat()}\" />
  <meta name=\"description\" content=\"تقرير مقارنة الجرد الشهري لمصنع الذهب\" />
  <meta name=\"format-detection\" content=\"telephone=no\" />
  <meta http-equiv=\"X-UA-Compatible\" content=\"IE=edge\" />
  <meta http-equiv=\"Content-Language\" content=\"ar\" />
  <meta name=\"color-scheme\" content=\"light only\" />
  <meta name=\"referrer\" content=\"no-referrer\" />
  <link rel=\"preconnect\" href=\"https://fonts.googleapis.com\" crossorigin>
  <link rel=\"preconnect\" href=\"https://fonts.gstatic.com\" crossorigin>
  <link href=\"https://fonts.googleapis.com/css2?family=Noto+Naskh+Arabic:wght@400;600&display=swap\" rel=\"stylesheet\">
  <script>window.onload = function(){{ window.print && window.print(); }}</script>
  </head>
<body>
  <div class=\"container\">
    <h1>{title}</h1>
    <div class=\"subtitle\">{subtitle}</div>
    {table_html}
  </div>
</body>
</html>
"""

    @staticmethod
    def _build_html_table(rows: List[Dict[str, object]], year: int, month: int) -> str:
        head = """
<table>
  <thead>
    <tr>
      <th>الصنف</th>
      <th>الوحدة</th>
      <th>كمية الشهر الحالي</th>
      <th>كمية الشهر السابق</th>
      <th>الفرق</th>
    </tr>
  </thead>
  <tbody>
"""
        body_rows = []
        for r in rows:
            current_qty = float(r.get("current_qty", 0) or 0)
            previous_qty = float(r.get("previous_qty", 0) or 0)
            diff = current_qty - previous_qty
            sign = "+" if diff > 0 else ("-" if diff < 0 else "")
            css_class = "inc" if diff > 0 else ("dec" if diff < 0 else "")
            body_rows.append(
                f"<tr>"
                f"<td>{r['item_name']}</td>"
                f"<td>{r['unit']}</td>"
                f"<td>{current_qty:,.3f}</td>"
                f"<td>{previous_qty:,.3f}</td>"
                f"<td class=\"{css_class}\">{sign}{abs(diff):,.3f}</td>"
                f"</tr>"
            )
        foot = """
  </tbody>
</table>
"""
        return head + "\n".join(body_rows) + foot

    @staticmethod
    def export_html(path: str, rows: List[Dict[str, object]], year: int, month: int) -> None:
        title = "تقرير المقارنة الشهري للجرد"
        subtitle = f"الشهر: {month:02d} / {year} — مصنع الذهب — تاريخ التقرير: {datetime.now():%Y-%m-%d %H:%M}"
        html = ExportManager._html_template(title, ExportManager._build_html_table(rows, year, month), subtitle)
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)

    @staticmethod
    def export_html_to_temp(rows: List[Dict[str, object]], year: int, month: int) -> str:
        tmp_dir = tempfile.gettempdir()
        path = os.path.join(tmp_dir, f"gold_inventory_report_{year}_{month:02d}.html")
        ExportManager.export_html(path, rows, year, month)
        return path

    @staticmethod
    def export_excel(path: str, rows: List[Dict[str, object]], year: int, month: int) -> None:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

        wb = Workbook()
        ws = wb.active
        ws.title = f"{year}-{month:02d}"

        # Title
        ws.merge_cells("A1:E1")
        ws["A1"] = "تقرير المقارنة الشهري للجرد — مصنع الذهب"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:E2")
        ws["A2"] = f"الشهر: {month:02d} / {year} — تاريخ التقرير: {datetime.now():%Y-%m-%d %H:%M}"
        ws["A2"].alignment = Alignment(horizontal="center")

        headers = ["الصنف", "الوحدة", "كمية الشهر الحالي", "كمية الشهر السابق", "الفرق"]
        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="ECEFF1")
        thin = Side(border_style="thin", color="D0D0D0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col in range(1, 6):
            cell = ws.cell(row=3, column=col)
            cell.fill = header_fill
            cell.border = border
            cell.font = Font(bold=True)

        inc_fill = PatternFill("solid", fgColor="90EE90")
        dec_fill = PatternFill("solid", fgColor="FFB6C1")

        for r in rows:
            current_qty = float(r.get("current_qty", 0) or 0)
            previous_qty = float(r.get("previous_qty", 0) or 0)
            diff = current_qty - previous_qty
            sign = "+" if diff > 0 else ("-" if diff < 0 else "")
            ws.append([
                r["item_name"],
                r["unit"],
                round(current_qty, 3),
                round(previous_qty, 3),
                f"{sign}{abs(round(diff, 3))}",
            ])

        # Apply borders and diff colors
        start_row = 4
        end_row = ws.max_row
        for row in range(start_row, end_row + 1):
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
            diff_value_str = str(ws.cell(row=row, column=5).value or "0")
            if diff_value_str.startswith("+"):
                ws.cell(row=row, column=5).fill = inc_fill
            elif diff_value_str.startswith("-"):
                ws.cell(row=row, column=5).fill = dec_fill

        # Column widths
        widths = [30, 12, 22, 22, 12]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(ord('A') + idx - 1)].width = width

        wb.save(path)

